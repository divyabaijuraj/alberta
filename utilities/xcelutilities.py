from openpyxl import load_workbook
def getRowCount(file,sheetName):
    print(sheetName,"sheet**")
    workbook = load_workbook(file)
    sheet = workbook[sheetName]
    #workbook=openpyxl.load_workbook(file)
    #sheet=workbook.get_sheet_by_name(sheetName)
    print("sheetssss:",sheet)
    return sheet.max_row

def getColumnCount(file,sheetName):
    workbook=load_workbook(file)
    #sheet=workbook.get_sheet_by_name(sheetName)
    sheet = workbook[sheetName]
    return sheet.max_column

def readData(file,sheetName,rownum,colnum):
    workbook=load_workbook(file)
    sheet = workbook[sheetName]
    user= sheet.cell(row=rownum,column=colnum).value
    return user

def writeData(file, sheetName, rownum, colnum,data):
    workbook = load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=rownum, column=colnum).value=data
    workbook.save(file)
